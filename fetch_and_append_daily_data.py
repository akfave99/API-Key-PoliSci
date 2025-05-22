import requests
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# Your Data.gov API key
data_gov_api_key = "pmGBCBa9Qfkawa1E1xhh7pxWkAN2kaboD4vK04n0"

# API endpoint for package_search
data_gov_url = "https://catalog.data.gov/api/3/action/package_search"

# Example: fetch the first 5 datasets (customize as needed)
params = {
    "rows": 5
}
headers = {
    "Authorization": data_gov_api_key
}

response = requests.get(data_gov_url, params=params, headers=headers)
response.raise_for_status()
data = response.json()

# Extract relevant info (customize fields as needed)
results = data["result"]["results"]
rows = []
for item in results:
    rows.append({
        "title": item.get("title"),
        "organization": item.get("organization", {}).get("title"),
        "metadata_created": item.get("metadata_created"),
        "retrieved_at": datetime.now().isoformat()
    })

# Convert to DataFrame
new_df = pd.DataFrame(rows)

# Path to your Excel file
excel_path = "/ak/desktop/daily_data.xslx"

# Append to Excel file
try:
    # Load existing workbook
    book = load_workbook(excel_path)
    writer = pd.ExcelWriter(excel_path, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    # Read existing data
    startrow = writer.sheets['Sheet1'].max_row
    new_df.to_excel(writer, index=False, header=False, startrow=startrow, sheet_name='Sheet1')
    writer.save()
    print("Data appended to existing Excel file.")
except FileNotFoundError:
    # If file doesn't exist, create it
    new_df.to_excel(excel_path, index=False)
    print("Excel file created and data written.")

# --- FEC API Integration ---
fec_api_key = data_gov_api_key  # Using the same API key
fec_base_url = "https://api.open.fec.gov/v1/candidates/search/"
fec_params = {
    "api_key": fec_api_key,
    "per_page": 5  # Fetch 5 candidates as an example
}
fec_response = requests.get(fec_base_url, params=fec_params)
fec_response.raise_for_status()
fec_data = fec_response.json()

# Extract relevant FEC candidate info
fec_results = fec_data.get("results", [])
fec_rows = []
for item in fec_results:
    fec_rows.append({
        "candidate_id": item.get("candidate_id"),
        "name": item.get("name"),
        "office_full": item.get("office_full"),
        "party_full": item.get("party_full"),
        "state": item.get("state"),
        "district": item.get("district"),
        "incumbent_challenge_full": item.get("incumbent_challenge_full"),
        "retrieved_at": datetime.now().isoformat()
    })

fec_df = pd.DataFrame(fec_rows)

# Append FEC data to a new sheet in the same Excel file
with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    fec_df.to_excel(writer, index=False, sheet_name='FEC_Candidates')
    print("FEC candidate data written to 'FEC_Candidates' sheet.")

# --- Census API Integration ---
census_api_key = "7c5962b163e1b18fa334f17b80a42fa5d0b361ff"
census_base_url = "https://api.census.gov/data/2020/dec/pl"
census_params = {
    "get": "NAME,P1_001N",  # Example: get place name and total population
    "for": "state:*",
    "key": census_api_key
}
census_response = requests.get(census_base_url, params=census_params)
census_response.raise_for_status()
census_data = census_response.json()

# Convert Census data to DataFrame
census_columns = census_data[0]
census_rows = census_data[1:]
census_df = pd.DataFrame(census_rows, columns=census_columns)
census_df["retrieved_at"] = datetime.now().isoformat()

# Append Census data to the existing sheet (Sheet1)
try:
    book = load_workbook(excel_path)
    writer = pd.ExcelWriter(excel_path, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    startrow = writer.sheets['Sheet1'].max_row
    census_df.to_excel(writer, index=False, header=False, startrow=startrow, sheet_name='Sheet1')
    writer.save()
    print("Census data appended to existing Excel file.")
except FileNotFoundError:
    census_df.to_excel(excel_path, index=False)
    print("Excel file created and Census data written.")
